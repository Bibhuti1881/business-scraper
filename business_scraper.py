import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import time

# ============================================================
#   BUSINESS SCRAPER — Nominatim Search API
#   No API key! Free! Works everywhere!
#   Built by Bibhuti Adhikari — bibhutiportfolio.vercel.app
# ============================================================

# ── CHANGE THESE ──
SEARCH_TERM = "cafe"        # cafe, restaurant, hotel, hospital, pharmacy, bank
CITY        = "Kathmandu"   # any city
MAX_RESULTS = 100

HEADERS = {
    "User-Agent": "BusinessScraper/1.0 bibhutiadhikari788@gmail.com",
    "Accept-Language": "en-US,en;q=0.9",
}

# ============================================================

def scrape(search_term, city, max_results):
    print(f"\n{'='*55}")
    print(f"  🏢 Business Scraper — Nominatim")
    print(f"  🔎 Search : {search_term} in {city}")
    print(f"  📦 Max    : {max_results}")
    print(f"{'='*55}\n")

    results = []
    page    = 0

    while len(results) < max_results:
        url    = "https://nominatim.openstreetmap.org/search"
        params = {
            "q"              : f"{search_term} in {city}",
            "format"         : "jsonv2",
            "addressdetails" : 1,
            "extratags"      : 1,
            "limit"          : 50,
            "offset"         : page * 50,
        }

        print(f"  ⏳ Fetching page {page+1}...")

        try:
            resp = requests.get(url, params=params, headers=HEADERS, timeout=15)
            data = resp.json()
        except Exception as e:
            print(f"  ❌ Error: {e}")
            break

        if not data:
            print(f"  ℹ️  No more results.")
            break

        for i, place in enumerate(data):
            name     = place.get("name") or place.get("display_name","").split(",")[0]
            addr     = place.get("address", {})
            extra    = place.get("extratags") or {}

            phone    = extra.get("phone") or extra.get("contact:phone") or "N/A"
            email    = extra.get("email") or extra.get("contact:email") or "N/A"
            website  = extra.get("website") or extra.get("contact:website") or "N/A"
            opening  = extra.get("opening_hours", "N/A")
            cuisine  = extra.get("cuisine", "N/A")
            category = place.get("type", search_term).replace("_"," ").title()

            # Build address
            parts = []
            for key in ["house_number","road","suburb","city","postcode","country"]:
                if addr.get(key): parts.append(addr[key])
            address = ", ".join(parts) if parts else place.get("display_name","N/A")

            idx = len(results) + 1
            results.append({
                "No."          : idx,
                "Business Name": name,
                "Phone"        : phone,
                "Email"        : email,
                "Address"      : address,
                "Category"     : category,
                "Cuisine/Type" : cuisine,
                "Opening Hours": opening,
                "Website"      : website,
            })
            print(f"  [{idx:03d}] {name[:40]:<40} | {phone}")

            if len(results) >= max_results:
                break

        page += 1
        time.sleep(1)  # be polite

        if len(data) < 50:
            break

    print(f"\n  ✅ Total scraped: {len(results)} businesses")
    return results


def export_excel(data, search_term, city):
    if not data:
        print("\n  ⚠️  No data!")
        return

    filename = f"businesses_{search_term}_{city}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb       = Workbook()
    ws       = wb.active
    ws.title = "Businesses"

    for i, w in enumerate([5,35,18,28,45,15,15,22,35], 1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w

    with_phone   = sum(1 for b in data if b["Phone"]   != "N/A")
    with_email   = sum(1 for b in data if b["Email"]   != "N/A")
    with_website = sum(1 for b in data if b["Website"] != "N/A")

    # Title
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = f"🏢 {search_term.title()}s in {city} — {datetime.now().strftime('%B %d, %Y')}"
    t.font      = Font(bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="1A1A2E")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle
    ws.merge_cells("A2:I2")
    s = ws["A2"]
    s.value     = f"Total: {len(data)} | Phones: {with_phone} | Emails: {with_email} | Websites: {with_website} | Source: OpenStreetMap | By: Bibhuti Adhikari"
    s.font      = Font(bold=True, size=10, color="FFFFFF")
    s.fill      = PatternFill("solid", fgColor="16213E")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = ["#","Business Name","Phone","Email","Address","Category","Cuisine/Type","Opening Hours","Website"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor="E55934")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, biz in enumerate(data, 4):
        fg   = "FFF5F2" if i % 2 == 0 else "FFFFFF"
        vals = [biz["No."],biz["Business Name"],biz["Phone"],biz["Email"],
                biz["Address"],biz["Category"],biz["Cuisine/Type"],
                biz["Opening Hours"],biz["Website"]]
        for col, val in enumerate(vals, 1):
            c           = ws.cell(i, col, val)
            c.fill      = PatternFill("solid", fgColor=fg)
            c.border    = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1: c.alignment = Alignment(horizontal="center", vertical="center")
            if col == 9: c.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[i].height = 16

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30
    ws2.merge_cells("A1:B1")
    h = ws2["A1"]
    h.value     = "📊 Scrape Summary"
    h.font      = Font(bold=True, size=13, color="FFFFFF")
    h.fill      = PatternFill("solid", fgColor="1A1A2E")
    h.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 28

    for row, (k, v) in enumerate([
        ("Total Businesses", len(data)),
        ("With Phone",       with_phone),
        ("With Email",       with_email),
        ("With Website",     with_website),
        ("Search Term",      search_term.title()),
        ("City",             city),
        ("Source",           "OpenStreetMap / Nominatim"),
        ("Scraped On",       datetime.now().strftime("%B %d, %Y")),
        ("Built by",         "Bibhuti Adhikari"),
        ("Portfolio",        "bibhutiportfolio.vercel.app"),
    ], 2):
        fg = "FFF5F2" if row % 2 == 0 else "FFFFFF"
        ws2[f"A{row}"] = k
        ws2[f"B{row}"] = str(v)
        ws2[f"A{row}"].font = Font(bold=True)
        ws2[f"A{row}"].fill = PatternFill("solid", fgColor=fg)
        ws2[f"B{row}"].fill = PatternFill("solid", fgColor=fg)

    wb.save(filename)
    print(f"\n  ✅ Excel saved → {filename}")
    print(f"  📊 {len(data)} businesses | {with_phone} phones | {with_email} emails | {with_website} websites")
    return filename


if __name__ == "__main__":
    print("🚀 Business Scraper — Nominatim (OpenStreetMap)")
    print("   Built by Bibhuti Adhikari")
    print("   bibhutiportfolio.vercel.app\n")

    data = scrape(SEARCH_TERM, CITY, MAX_RESULTS)

    if data:
        export_excel(data, SEARCH_TERM, CITY)
        print("\n🎉 Done! Open Excel to see your data.")
    else:
        print("\n❌ No results found.")